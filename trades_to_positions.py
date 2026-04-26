#!/usr/bin/env python3
"""
trades_to_positions.py

Convert a TradingView strategy "List of trades" export (.xlsx) into a
position-based summary, where consecutive entries that overlap in time
(e.g. an initial Fade EXT entry plus its Pyramid 2/3/4 adds) are rolled
up into a single position.

A new position starts every time the running contract size goes from
flat to non-flat; it closes when the running size returns to flat.

The output Positions sheet uses Excel row grouping: each position is a
parent (summary) row, with one collapsible child row per constituent
trade pair (entry + exit). Click the +/- in the outline gutter (or use
the 1/2 buttons at the top-left) to collapse or expand all children.

Usage:
    python trades_to_positions.py INPUT.xlsx [OUTPUT.xlsx] [--print-summary]

If OUTPUT is omitted, the file is written next to INPUT with suffix
"_positions.xlsx".

Tested against TradingView strategy exports for the Mean Reversion
Extension-Fade strategy (NQ/MNQ futures), but works on any TV strategy
export that uses the standard "List of trades" sheet schema.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula


REQUIRED_COLS = [
    "Trade #",
    "Type",
    "Date and time",
    "Signal",
    "Price USD",
    "Size (qty)",
    "Net P&L USD",
    "Net P&L %",
    "Favorable excursion USD",
    "Adverse excursion USD",
]

MONEY_FMT = '$#,##0.00;($#,##0.00);-'
PCT_FMT = '0.0%;(0.0%);-'
PCT_TV_FMT = '0.00"%";(0.00"%");-'  # TV stores % as "percentage points" already
DT_FMT = 'yyyy-mm-dd hh:mm'


# ---------------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------------

def load_trades(input_path: Path) -> pd.DataFrame:
    """Load the 'List of trades' sheet and validate required columns."""
    df = pd.read_excel(input_path, sheet_name="List of trades")
    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        raise ValueError(
            f"Input file is missing required columns: {missing}. "
            f"Expected a TradingView strategy 'List of trades' export."
        )
    return df


def aggregate_positions(df: pd.DataFrame) -> tuple[pd.DataFrame, list[list[dict]]]:
    """
    Walk the trade rows in chronological order, tracking the running
    contract size. Each transition from flat to non-flat starts a new
    position; each return to flat closes it.

    Returns:
        positions_df: one summary row per position
        children:     parallel list; children[i] is a list of trade-pair
                      dicts (entry+exit) that belong to position i+1
    """
    df = df.copy()
    df["is_entry"] = df["Type"].str.startswith("Entry")
    df["signed_qty"] = df.apply(
        lambda r: (1 if "long" in r["Type"] else -1)
        * r["Size (qty)"]
        * (1 if r["is_entry"] else -1),
        axis=1,
    )
    # At identical timestamps within the same Trade #, entries come
    # before exits (a same-bar open/close has the open first).
    df["_entry_first"] = (~df["is_entry"]).astype(int)
    df = df.sort_values(
        by=["Date and time", "Trade #", "_entry_first"], kind="stable"
    ).reset_index(drop=True)

    positions: list[dict] = []
    current: dict | None = None
    running_size = 0

    for _, row in df.iterrows():
        prev_size = running_size
        running_size += row["signed_qty"]

        if row["is_entry"]:
            if prev_size == 0:
                current = {
                    "entries": [],
                    "exits": [],
                    "direction": "Long" if "long" in row["Type"] else "Short",
                }
            current["entries"].append(row)
        else:
            if current is None:
                continue
            current["exits"].append(row)
            if running_size == 0:
                positions.append(current)
                current = None

    if running_size != 0:
        print(
            f"WARNING: trade list ended with non-zero running size ({running_size}). "
            "Last position may be incomplete.",
            file=sys.stderr,
        )

    summary_rows = []
    children_rows: list[list[dict]] = []

    for idx, p in enumerate(positions, 1):
        entries_df = pd.DataFrame(p["entries"])
        exits_df = pd.DataFrame(p["exits"])
        total_qty = int(entries_df["Size (qty)"].sum())

        entry_time = entries_df["Date and time"].min()
        last_add_time = entries_df["Date and time"].max()
        exit_time = exits_df["Date and time"].max()

        avg_entry = (entries_df["Price USD"] * entries_df["Size (qty)"]).sum() / total_qty
        avg_exit = (exits_df["Price USD"] * exits_df["Size (qty)"]).sum() / total_qty

        # Net P&L is recorded on both the entry and the exit row of each
        # Trade #; summing entries gives the position's realized P&L
        # without double-counting.
        total_pnl = entries_df["Net P&L USD"].sum()
        total_pnl_pct = entries_df["Net P&L %"].sum()
        duration_min = (exit_time - entry_time).total_seconds() / 60

        initial_signal = entries_df.sort_values("Date and time").iloc[0]["Signal"]
        final_signal = exits_df.sort_values("Date and time").iloc[-1]["Signal"]
        mfe = entries_df["Favorable excursion USD"].max()
        mae = entries_df["Adverse excursion USD"].min()

        trade_ids = sorted(entries_df["Trade #"].astype(int).unique().tolist())

        summary_rows.append({
            "Position #": idx,
            "Trade #": "",
            "Direction": p["direction"],
            "Entry signal": initial_signal,
            "Entry time": entry_time,
            "Entry price": round(avg_entry, 4),
            "Exit signal": final_signal,
            "Exit time": exit_time,
            "Exit price": round(avg_exit, 4),
            "Qty": total_qty,
            "# Entries": len(entries_df),
            "Net P&L USD": round(total_pnl, 2),
            "Net P&L %": round(total_pnl_pct, 4),
            "MFE USD": round(mfe, 2),
            "MAE USD": round(mae, 2),
            "Duration (min)": round(duration_min, 1),
            "Trade #s": ", ".join(str(t) for t in trade_ids),
        })

        # Build child rows (one per constituent trade pair)
        children: list[dict] = []
        for tid in trade_ids:
            ent = entries_df[entries_df["Trade #"] == tid].iloc[0]
            ext = exits_df[exits_df["Trade #"] == tid].iloc[0]
            child_dur = (ext["Date and time"] - ent["Date and time"]).total_seconds() / 60
            children.append({
                "Position #": "",
                "Trade #": int(tid),
                "Direction": p["direction"],
                "Entry signal": ent["Signal"],
                "Entry time": ent["Date and time"],
                "Entry price": ent["Price USD"],
                "Exit signal": ext["Signal"],
                "Exit time": ext["Date and time"],
                "Exit price": ext["Price USD"],
                "Qty": int(ent["Size (qty)"]),
                "# Entries": "",
                "Net P&L USD": round(ent["Net P&L USD"], 2),
                "Net P&L %": round(ent["Net P&L %"], 4),
                "MFE USD": round(ent["Favorable excursion USD"], 2),
                "MAE USD": round(ent["Adverse excursion USD"], 2),
                "Duration (min)": round(child_dur, 1),
                "Trade #s": "",
            })
        children_rows.append(children)

    pos_df = pd.DataFrame(summary_rows)
    if not pos_df.empty:
        pos_df["Cumulative P&L USD"] = pos_df["Net P&L USD"].cumsum().round(2)
    return pos_df, children_rows


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

POS_HEADERS = [
    "Position #",
    "Trade #",
    "Direction",
    "Entry signal",
    "Entry time",
    "Entry price",
    "Exit signal",
    "Exit time",
    "Exit price",
    "Qty",
    "# Entries",
    "Net P&L USD",
    "Net P&L %",
    "MFE USD",
    "MAE USD",
    "Duration (min)",
    "Trade #s",
    "Cumulative P&L USD",
]

WIDTHS = {
    "Position #": 11,
    "Trade #": 9,
    "Direction": 10,
    "Entry signal": 14,
    "Entry time": 18,
    "Entry price": 12,
    "Exit signal": 16,
    "Exit time": 18,
    "Exit price": 12,
    "Qty": 7,
    "# Entries": 11,
    "Net P&L USD": 13,
    "Net P&L %": 11,
    "MFE USD": 12,
    "MAE USD": 12,
    "Duration (min)": 13,
    "Trade #s": 16,
    "Cumulative P&L USD": 16,
}

NUM_FMTS = {
    "Position #": "0",
    "Trade #": "0",
    "Entry time": DT_FMT,
    "Entry price": "#,##0.00",
    "Exit time": DT_FMT,
    "Exit price": "#,##0.00",
    "Qty": "0",
    "# Entries": "0",
    "Net P&L USD": MONEY_FMT,
    "Net P&L %": PCT_TV_FMT,
    "MFE USD": MONEY_FMT,
    "MAE USD": MONEY_FMT,
    "Duration (min)": "0",
    "Cumulative P&L USD": MONEY_FMT,
}


def _write_cell(ws, row: int, col: int, value, *, font, fill=None) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fill is not None:
        c.fill = fill
    c.border = BORDER
    c.alignment = Alignment(vertical="center")


def _write_positions_sheet(
    wb: Workbook, pos_df: pd.DataFrame, children: list[list[dict]]
) -> tuple[int, list[int], list[tuple[int, int]]]:
    """
    Write the Positions sheet with hierarchical parent/child rows.
    Returns (n_positions, parent_row_indices, child_ranges) where
    parent_row_indices[i] is the Excel row of position i+1 and
    child_ranges[i] = (first_child_row, last_child_row) for that position.
    """
    ws = wb.active
    ws.title = "Positions"

    # Outline: parent (summary) rows ABOVE their children, gutter on the left
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False

    # Header row
    ws.append(POS_HEADERS)
    for col_idx in range(1, len(POS_HEADERS) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = BORDER
    ws.row_dimensions[1].height = 38
    ws.freeze_panes = "A2"

    # Column widths
    for h, w in WIDTHS.items():
        ws.column_dimensions[get_column_letter(POS_HEADERS.index(h) + 1)].width = w

    parent_row_idx_for: dict[int, int] = {}
    child_ranges: list[tuple[int, int]] = []
    parent_rows: list[int] = []

    parent_font = Font(name="Arial", bold=True, size=10)
    parent_fill = PatternFill("solid", start_color="F2F2F2")
    child_font = Font(name="Arial", size=9, color="595959")
    child_fill_long = PatternFill("solid", start_color="EAF1FB")
    child_fill_short = PatternFill("solid", start_color="FDEDE7")
    long_color = Font(name="Arial", size=10, color="1F4E78", bold=True)
    short_color = Font(name="Arial", size=10, color="C00000", bold=True)

    cur_row = 2
    for i, parent_record in enumerate(pos_df.to_dict(orient="records")):
        # ---- Parent (position summary) row ----
        for col_idx, h in enumerate(POS_HEADERS, 1):
            val = parent_record.get(h, "")
            _write_cell(ws, cur_row, col_idx, val, font=parent_font, fill=parent_fill)

        # Direction tint on the parent's Direction cell
        dir_col = POS_HEADERS.index("Direction") + 1
        dir_cell = ws.cell(row=cur_row, column=dir_col)
        if parent_record["Direction"] == "Long":
            dir_cell.fill = PatternFill("solid", start_color="D9E1F2")
            dir_cell.font = long_color
        else:
            dir_cell.fill = PatternFill("solid", start_color="FCE4D6")
            dir_cell.font = short_color

        parent_rows.append(cur_row)
        parent_row_idx_for[i + 1] = cur_row
        cur_row += 1

        # ---- Children (constituent trades) ----
        first_child = cur_row
        for child in children[i]:
            child_fill = child_fill_long if parent_record["Direction"] == "Long" else child_fill_short
            for col_idx, h in enumerate(POS_HEADERS, 1):
                val = child.get(h, "")
                _write_cell(ws, cur_row, col_idx, val, font=child_font, fill=child_fill)
            # Indent visual cue: prefix Trade # cell with a small marker
            tnum_col = POS_HEADERS.index("Trade #") + 1
            tcell = ws.cell(row=cur_row, column=tnum_col)
            tcell.alignment = Alignment(vertical="center", indent=1)
            # Mark this row as outline level 1 so it groups under the parent
            ws.row_dimensions[cur_row].outline_level = 1
            cur_row += 1
        last_child = cur_row - 1
        child_ranges.append((first_child, last_child))

    # Number formats — apply across all data rows
    for h, fmt in NUM_FMTS.items():
        col = POS_HEADERS.index(h) + 1
        for r in range(2, cur_row):
            ws.cell(row=r, column=col).number_format = fmt

    n = len(pos_df)

    # Cumulative P&L formula on parent rows only (column R)
    pnl_col = POS_HEADERS.index("Net P&L USD") + 1
    cum_col = POS_HEADERS.index("Cumulative P&L USD") + 1
    pnl_letter = get_column_letter(pnl_col)
    cum_letter = get_column_letter(cum_col)
    for i, parent_row in enumerate(parent_rows):
        if i == 0:
            ws.cell(row=parent_row, column=cum_col, value=f"={pnl_letter}{parent_row}")
        else:
            prev_parent = parent_rows[i - 1]
            ws.cell(
                row=parent_row,
                column=cum_col,
                value=f"={cum_letter}{prev_parent}+{pnl_letter}{parent_row}",
            )

    # Conditional fill on Net P&L for parent rows
    parent_pnl_addresses = ",".join(
        f"{pnl_letter}{r}" for r in parent_rows
    )
    # CellIsRule supports a single contiguous range; we'd need multiple
    # rules to cover non-contiguous parent rows. Easier: apply across the
    # full data column — children get a subtle tint anyway and parents
    # still pop with bold + grey background.
    if parent_rows:
        full_range = f"{pnl_letter}2:{pnl_letter}{cur_row-1}"
        ws.conditional_formatting.add(
            full_range,
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                font=Font(name="Arial", size=10, color="006100"),
            ),
        )
        ws.conditional_formatting.add(
            full_range,
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                font=Font(name="Arial", size=10, color="9C0006"),
            ),
        )

    return n, parent_rows, child_ranges


def _write_summary_sheet(
    wb: Workbook, pos_df: pd.DataFrame, parent_rows: list[int]
) -> None:
    """Summary sheet using parent-row references in the Positions sheet."""
    ws = wb.create_sheet("Summary")
    for col, w in zip("ABCD", (36, 16, 16, 16)):
        ws.column_dimensions[col].width = w

    def section(row: int, text: str) -> None:
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 22

    def col_hdr(row: int, cols: list[str]) -> None:
        for i, t in enumerate(cols, 1):
            c = ws.cell(row=row, column=i, value=t)
            c.font = Font(name="Arial", bold=True, size=10)
            c.fill = PatternFill("solid", start_color="D9E1F2")
            c.alignment = Alignment(horizontal="center")
            c.border = BORDER

    title = ws.cell(row=1, column=1, value="Position-Based Performance Summary")
    title.font = Font(name="Arial", bold=True, size=14)
    ws.merge_cells("A1:D1")

    # Because parent rows are non-contiguous in the Positions sheet, we
    # compute summary stats directly in Python and write the values; the
    # totals section still uses formulas where they make sense.
    n = len(pos_df)
    longs = pos_df[pos_df["Direction"] == "Long"]
    shorts = pos_df[pos_df["Direction"] == "Short"]

    section(3, "Overall")
    col_hdr(4, ["Metric", "All", "Long", "Short"])

    def write_row(r: int, label: str, all_v, long_v, short_v) -> None:
        ws.cell(row=r, column=1, value=label).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=2, value=all_v)
        ws.cell(row=r, column=3, value=long_v)
        ws.cell(row=r, column=4, value=short_v)
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BORDER

    wins = pos_df[pos_df["Net P&L USD"] > 0]
    losses = pos_df[pos_df["Net P&L USD"] < 0]
    breakeven = pos_df[pos_df["Net P&L USD"] == 0]

    rows = [
        ("Total positions", n, len(longs), len(shorts), "0"),
        ("Winning positions", len(wins),
         len(wins[wins["Direction"] == "Long"]),
         len(wins[wins["Direction"] == "Short"]), "0"),
        ("Losing positions", len(losses),
         len(losses[losses["Direction"] == "Long"]),
         len(losses[losses["Direction"] == "Short"]), "0"),
        ("Breakeven positions", len(breakeven),
         len(breakeven[breakeven["Direction"] == "Long"]),
         len(breakeven[breakeven["Direction"] == "Short"]), "0"),
        ("Win rate",
         len(wins) / n if n else 0,
         len(wins[wins["Direction"] == "Long"]) / len(longs) if len(longs) else 0,
         len(wins[wins["Direction"] == "Short"]) / len(shorts) if len(shorts) else 0,
         PCT_FMT),
        ("Total Net P&L (USD)",
         round(pos_df["Net P&L USD"].sum(), 2),
         round(longs["Net P&L USD"].sum(), 2),
         round(shorts["Net P&L USD"].sum(), 2),
         MONEY_FMT),
        ("Avg position P&L (USD)",
         round(pos_df["Net P&L USD"].mean(), 2) if n else 0,
         round(longs["Net P&L USD"].mean(), 2) if len(longs) else 0,
         round(shorts["Net P&L USD"].mean(), 2) if len(shorts) else 0,
         MONEY_FMT),
        ("Best position (USD)",
         round(pos_df["Net P&L USD"].max(), 2),
         round(longs["Net P&L USD"].max(), 2) if len(longs) else 0,
         round(shorts["Net P&L USD"].max(), 2) if len(shorts) else 0,
         MONEY_FMT),
        ("Worst position (USD)",
         round(pos_df["Net P&L USD"].min(), 2),
         round(longs["Net P&L USD"].min(), 2) if len(longs) else 0,
         round(shorts["Net P&L USD"].min(), 2) if len(shorts) else 0,
         MONEY_FMT),
        ("Avg # entries / position",
         round(pos_df["# Entries"].mean(), 2),
         round(longs["# Entries"].mean(), 2) if len(longs) else 0,
         round(shorts["# Entries"].mean(), 2) if len(shorts) else 0,
         "0.00"),
        ("Pyramided positions (>1 entry)",
         int((pos_df["# Entries"] > 1).sum()),
         int((longs["# Entries"] > 1).sum()),
         int((shorts["# Entries"] > 1).sum()),
         "0"),
        ("Avg duration (min)",
         round(pos_df["Duration (min)"].mean(), 0),
         round(longs["Duration (min)"].mean(), 0) if len(longs) else 0,
         round(shorts["Duration (min)"].mean(), 0) if len(shorts) else 0,
         "0"),
    ]
    for i, (label, all_v, long_v, short_v, fmt) in enumerate(rows, start=5):
        write_row(i, label, all_v, long_v, short_v)
        for col in (2, 3, 4):
            ws.cell(row=i, column=col).number_format = fmt

    # Pyramid breakdown
    pyr_start = 5 + len(rows) + 1
    section(pyr_start, "By # Entries (Pyramid Levels)")
    col_hdr(pyr_start + 1, ["# Entries", "Count", "Total P&L", "Avg P&L"])
    grp = pos_df.groupby("# Entries")["Net P&L USD"].agg(["count", "sum", "mean"])
    for j, (lvl, row) in enumerate(grp.iterrows()):
        r = pyr_start + 2 + j
        ws.cell(row=r, column=1, value=int(lvl))
        ws.cell(row=r, column=2, value=int(row["count"]))
        ws.cell(row=r, column=3, value=round(row["sum"], 2))
        ws.cell(row=r, column=4, value=round(row["mean"], 2))
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BORDER
        ws.cell(row=r, column=2).number_format = "0"
        ws.cell(row=r, column=3).number_format = MONEY_FMT
        ws.cell(row=r, column=4).number_format = MONEY_FMT

    # Exit-reason breakdown
    exit_start = pyr_start + 2 + len(grp) + 1
    section(exit_start, "By Final Exit Signal")
    col_hdr(exit_start + 1, ["Exit signal", "Count", "Total P&L", "Avg P&L"])
    # Use the parent-row "Exit signal" column for the breakdown
    egrp = pos_df.groupby("Exit signal")["Net P&L USD"].agg(["count", "sum", "mean"])
    for j, (sig, row) in enumerate(egrp.iterrows()):
        r = exit_start + 2 + j
        ws.cell(row=r, column=1, value=sig)
        ws.cell(row=r, column=2, value=int(row["count"]))
        ws.cell(row=r, column=3, value=round(row["sum"], 2))
        ws.cell(row=r, column=4, value=round(row["mean"], 2))
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BORDER
        ws.cell(row=r, column=2).number_format = "0"
        ws.cell(row=r, column=3).number_format = MONEY_FMT
        ws.cell(row=r, column=4).number_format = MONEY_FMT


def _write_trades_sheet(wb: Workbook, trades_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("List of trades")
    headers = list(trades_df.columns)
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = BORDER
    for row in trades_df.itertuples(index=False):
        ws.append(list(row))
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    if "Date and time" in headers:
        ws.column_dimensions[get_column_letter(headers.index("Date and time") + 1)].width = 18
    if "Signal" in headers:
        ws.column_dimensions[get_column_letter(headers.index("Signal") + 1)].width = 14
    if "Type" in headers:
        ws.column_dimensions[get_column_letter(headers.index("Type") + 1)].width = 12
    ws.freeze_panes = "A2"
    money_cols = [
        "Price USD",
        "Size (value)",
        "Net P&L USD",
        "Favorable excursion USD",
        "Adverse excursion USD",
        "Cumulative P&L USD",
    ]
    for r in range(2, ws.max_row + 1):
        if "Date and time" in headers:
            ws.cell(row=r, column=headers.index("Date and time") + 1).number_format = DT_FMT
        for col_name in money_cols:
            if col_name in headers:
                ws.cell(row=r, column=headers.index(col_name) + 1).number_format = MONEY_FMT


def write_workbook(
    output_path: Path,
    pos_df: pd.DataFrame,
    children: list[list[dict]],
    trades_df: pd.DataFrame,
) -> None:
    if pos_df.empty:
        raise ValueError("No positions found in trade list — nothing to write.")
    wb = Workbook()
    n, parent_rows, _ = _write_positions_sheet(wb, pos_df, children)
    _write_summary_sheet(wb, pos_df, parent_rows)
    _write_trades_sheet(wb, trades_df)
    wb.save(output_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Convert TradingView trade list to position-aggregated xlsx."
    )
    parser.add_argument("input", type=Path, help="Path to TV trade-list .xlsx")
    parser.add_argument(
        "output",
        type=Path,
        nargs="?",
        default=None,
        help="Output .xlsx path (default: <input>_positions.xlsx)",
    )
    parser.add_argument(
        "--print-summary",
        action="store_true",
        help="Print a text summary of positions to stdout",
    )
    args = parser.parse_args(argv)

    if not args.input.exists():
        print(f"Input file not found: {args.input}", file=sys.stderr)
        return 1

    output = args.output or args.input.with_name(
        args.input.stem + "_positions.xlsx"
    )

    print(f"Reading: {args.input}")
    trades_df = load_trades(args.input)
    print(f"  {len(trades_df)} trade rows ({trades_df['Trade #'].nunique()} trades)")

    pos_df, children = aggregate_positions(trades_df)
    print(f"  Aggregated into {len(pos_df)} positions")

    write_workbook(output, pos_df, children, trades_df)
    print(f"Wrote: {output}")

    if args.print_summary:
        print()
        print(_text_summary(pos_df))

    return 0


def _text_summary(pos_df: pd.DataFrame) -> str:
    if pos_df.empty:
        return "(no positions)"
    lines = []
    total = len(pos_df)
    longs = pos_df[pos_df["Direction"] == "Long"]
    shorts = pos_df[pos_df["Direction"] == "Short"]
    wins = pos_df[pos_df["Net P&L USD"] > 0]

    def fmt_money(v: float) -> str:
        return f"${v:,.2f}" if v >= 0 else f"(${abs(v):,.2f})"

    lines.append(f"Positions: {total}  (Long {len(longs)} / Short {len(shorts)})")
    lines.append(
        f"Win rate:  {len(wins)/total:.1%}  "
        f"(Long {len(wins[wins['Direction']=='Long'])/max(len(longs),1):.1%}, "
        f"Short {len(wins[wins['Direction']=='Short'])/max(len(shorts),1):.1%})"
    )
    lines.append(
        f"Net P&L:   {fmt_money(pos_df['Net P&L USD'].sum())}  "
        f"(Long {fmt_money(longs['Net P&L USD'].sum())}, "
        f"Short {fmt_money(shorts['Net P&L USD'].sum())})"
    )
    lines.append(
        f"Avg P&L:   {fmt_money(pos_df['Net P&L USD'].mean())}  |  "
        f"Best {fmt_money(pos_df['Net P&L USD'].max())}  |  "
        f"Worst {fmt_money(pos_df['Net P&L USD'].min())}"
    )
    pyramided = pos_df[pos_df["# Entries"] > 1]
    lines.append(
        f"Pyramided: {len(pyramided)} of {total} "
        f"(avg {pos_df['# Entries'].mean():.2f} entries/position)"
    )
    lines.append("")
    lines.append("By # Entries:")
    grp = pos_df.groupby("# Entries").agg(
        count=("Net P&L USD", "size"),
        total_pnl=("Net P&L USD", "sum"),
        avg_pnl=("Net P&L USD", "mean"),
    )
    for n, row in grp.iterrows():
        lines.append(
            f"  {int(n)}: {int(row['count']):>3} positions, "
            f"total {fmt_money(row['total_pnl']):>14}, "
            f"avg {fmt_money(row['avg_pnl']):>12}"
        )
    return "\n".join(lines)


if __name__ == "__main__":
    sys.exit(main())
